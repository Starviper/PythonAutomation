import openpyxl

wb = openpyxl.load_workbook('C:/Special/PythonProj/example.xlsx')
sheets = wb.sheetnames
print(type(sheets))
print(sheets[1].title())

Mainsheet = wb.active
print(Mainsheet)

cellB = Mainsheet['B1']
print('Row %s Column %s is %s' % (cellB.row, cellB.column, cellB.value))

for i in range(1, 8, 2):
    print(i, Mainsheet.cell(row=i, column=2).value)

print('Max rows:', Mainsheet.max_row, 'Max columns:', Mainsheet.max_column)


# for rowOfCellObjects in tuple(Mainsheet['A1':'C3']):
#    for cellObj in tuple(Mainsheet['A1':'C3']):
#        print(cellObj.coordinate, cellObj.value)
#    print('--- END OF ROW ---')
