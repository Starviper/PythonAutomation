#!python3

# VariousExcel.py - various settings that can be easily changed with Excel documents via openpyxl

import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active

italic24Font = Font(size=12, italic=True) # Create a Font.
sheet['A1'].font = italic24Font # Apply the font to A1
sheet['A1'] = 'FontTest'

TNR = Font(name='Times New Roman', bold=True)
sheet['A2'].font = TNR
sheet['A2'] = 'Times'

sheet['A3'] = 200
sheet['A4'] = 300
sheet['A5'] = '=SUM(A1:A2)' # Formula test

for i in range(1, 11): # Create some junk data in column C
    sheet['C' + str(i)] = i

# Create a bar chart
refObj = openpyxl.chart.Reference(sheet, min_col=3, min_row=1, max_col=3, max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title='Series test')
chartObj = openpyxl.chart.BarChart()
chartObj.title = 'Chart Test'
chartObj.append(seriesObj)

sheet.add_chart(chartObj, 'D3')

# Height / Width -
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20

#Cell merge
sheet.merge_cells('B2:B4')
# sheet.unmerge_cells('B2:B4')

sheet.freeze_panes = 'C3' # Row 1/2 and columns A/B will be 'frozen'

# Save the document
wb.save('examples.xlsx')